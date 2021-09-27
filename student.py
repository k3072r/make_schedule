import openpyxl as px
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Color, colors
from openpyxl.styles.fills import PatternFill

def get_students(wb):

    ws_student = wb["生徒スケジュール"]
    ws_register = wb["時間割登録"]

    # 講習期間の授業日数を取得
    count = 0
    for row in ws_register.iter_rows(min_row=6, min_col=2):
        if row[0].value is None:
            break
        count += 1

    lastcolumn = ws_student.max_column

    students = []

    for row in ws_student.iter_rows(min_row=4, min_col=4):

        name = row[0].value
        if name is None:
            break

        
        # 生徒スケジュールを以下のような配列に格納
        # students = [
        #               [name, [[1日目のスケジュール（コマを表す要素数7の配列）], [2日目のスケジュール], ...]],
        #               [name, ...],
        #               ... ,
        #            ]
        #
        # 授業のコマは科目名、白抜きコマはfree、色付きコマはlockをそれぞれ文字列として格納
        student = [name]
        days = []
        day = []
        koma_count = 0
        for cell in row[2: 7 * count + 3]:

            # 配列の要素が7になったら次の日へ
            if koma_count >= 7:
                days.append(day)
                day = []
                koma_count = 0

            # 結合セルはlock扱い（lockでしか結合セルは使わないと思うので）
            if isinstance(cell, MergedCell):
                day.append("lock")

            elif cell.fill == PatternFill(fill_type=None):
                if cell.value is None:
                    day.append("free")
                else:
                    day.append(cell.value)
            
            else:
                day.append("lock")
            
            koma_count += 1

        student.append(days)
        students.append(student)

    return students




def student_free_count(students):
    free_counts = []
    for student in students:
        count = 0
        for day in student[1]:
            for lesson in day:
                if lesson == "free":
                    count += 1
        free_count = [student[0], count]
        free_counts.append(free_count)
    return free_counts

def free_count_bubble_sort(free_counts):
    length = len(free_counts)
    for i in range(0, length - 1):
        min = free_counts[i][1]
        indice = i
        for j in range(i + 1, length):
            if free_counts[j][1] < min:
                min = free_counts[j][1]
                indice = j
        free_counts[i], free_counts[indice] = free_counts[indice], free_counts[i]
        #hoge = free_counts[indice]
        #free_counts[indice] = free_counts[i]
        #free_counts[i] = hoge
    return free_counts