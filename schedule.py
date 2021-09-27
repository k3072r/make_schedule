import openpyxl as px
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Color, colors
from openpyxl.styles.fills import PatternFill

def cell_to_string(cell):
    if cell.fill == PatternFill(fill_type=None):
        if cell.value is None:
            return "free"
        else:
            return cell.value
    else:
        return "lock"

def get_schedule(wb):

    ws_schedule = wb["教室時間割"]
    ws_initial = wb["初期設定"]

    teachers = []
    for initial_row in ws_initial.iter_rows(min_row=5, min_col=8):
        name = initial_row[0].value
        teacher = [name]
        days = []

        # 該当講師が上から見て初めて現れる行番号を取得
        first_row = 4
        for row in ws_schedule.iter_rows(min_row=4, min_col=3):
            if name == row[0].value:
                break
            first_row += 1

        count = 0
        for row in ws_schedule.iter_rows(min_row=first_row, min_col=2):
            if row[0].value is None:
                break

            if count % 120 == 0:
                day = []
                for cell in row[2:21:3]:
                    #[[名前, 科目], [名前, 科目]]の形式でそのコマの内容を取得
                    upside_grade = cell_to_string(cell)
                    upside_name = cell_to_string(cell.offset(0,1))
                    upside_subj = cell_to_string(cell.offset(0,2))
                    downside_grade = cell_to_string(cell.offset(1,0))
                    downside_name = cell_to_string(cell.offset(1,1))
                    downside_subj = cell_to_string(cell.offset(1,2))
                    tuple = [[upside_grade, upside_name, upside_subj], [downside_grade, downside_name, downside_subj]]
                    day.append(tuple)    
                days.append(day)

            count += 1
        teacher.append(days)
        teachers.append(teacher)

    return teachers


# 教室時間割を以下の形式で取得
# teachers = [
#               [name, days],
#               [name, days],
#               ... ,
#            ]
# days = [day1, day2, ...]
# day = [lesson1, lesson2, ... , lesson7]
# lesson = [[grade, student, subject], [grade, student, subject]]