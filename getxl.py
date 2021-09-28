import openpyxl as px
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Color, colors
from openpyxl.styles.fills import PatternFill
from openpyxl import cell
from openpyxl.xml.constants import MIN_ROW


def get_students(wb):

    ws_student = wb["生徒スケジュール"]
    ws_register = wb["時間割登録"]

    # 講習期間の授業日数を取得
    count = 0
    for row in ws_register.iter_rows(min_row=6, min_col=2):
        if row[0].value is None:
            break
        count += 1

    students = []

    for row in ws_student.iter_rows(min_row=4, min_col=4):

        name = row[0].value
        if name is None:
            break
        
        # 生徒スケジュールを以下のような配列に格納
        # students = [
        #               [name, [[1日目のスケジュール（コマを表す要素数7の配列）], [2日目のスケジュール], ...]],
        #               [name, ...],
        #               ... ,
        #            ]
        #
        # 授業のコマは科目名、白抜きコマはfree、色付きコマはlockをそれぞれ文字列として格納
        student = [name]
        days = []
        day = []
        koma_count = 0
        for cell in row[2: 7 * count + 3]:

            # 配列の要素が7になったら次の日へ
            if koma_count >= 7:
                days.append(day)
                day = []
                koma_count = 0

            # 結合セルはlock扱い（lockでしか結合セルは使わないと思うので）
            if isinstance(cell, MergedCell):
                day.append("lock")

            elif cell.fill == PatternFill(fill_type=None):
                if cell.value is None:
                    day.append("free")
                else:
                    day.append(cell.value)
            
            else:
                day.append("lock")
            
            koma_count += 1

        student.append(days)
        students.append(student)

    return students




def cell_to_string(cell):
    if cell.fill == PatternFill(fill_type=None):
        if cell.value is None:
            return "free"
        else:
            return cell.value
    else:
        return "lock"

def get_schedule(wb):

    ws_schedule = wb["教室時間割"]
    ws_initial = wb["初期設定"]

    teachers = []
    for initial_row in ws_initial.iter_rows(min_row=5, min_col=8):
        name = initial_row[0].value
        teacher = [name]
        days = []

        # 該当講師が上から見て初めて現れる行番号を取得
        first_row = 4
        for row in ws_schedule.iter_rows(min_row=4, min_col=3):
            if name == row[0].value:
                break
            first_row += 1

        count = 0
        for row in ws_schedule.iter_rows(min_row=first_row, min_col=2):
            if row[0].value is None:
                break

            if count % 120 == 0:
                day = []
                for cell in row[2:21:3]:
                    #[[名前, 科目], [名前, 科目]]の形式でそのコマの内容を取得
                    upside_grade = cell_to_string(cell)
                    upside_name = cell_to_string(cell.offset(0,1))
                    upside_subj = cell_to_string(cell.offset(0,2))
                    downside_grade = cell_to_string(cell.offset(1,0))
                    downside_name = cell_to_string(cell.offset(1,1))
                    downside_subj = cell_to_string(cell.offset(1,2))
                    tuple = [[upside_grade, upside_name, upside_subj], [downside_grade, downside_name, downside_subj]]
                    day.append(tuple)    
                days.append(day)

            count += 1
        teacher.append(days)
        teachers.append(teacher)

    return teachers

# 教室時間割を以下の形式で取得
# teachers = [
#               [name, days],
#               [name, days],
#               ... ,
#            ]
# days = [day1, day2, ...]
# day = [lesson1, lesson2, ... , lesson7]
# lesson = [[grade, student, subject], [grade, student, subject]]




def get_lessons(wb):

    ws_hopes = wb["希望講師"]

    lessons = []

    for row in ws_hopes.iter_rows(min_row=4, min_col=2):

        grade = row[0].value
        name = row[2].value

        if name is None:
            break

        lesson = [grade, name]
        for cell in row[3:5]:
            lesson.append(cell.value)

        teachers = []
        for cell in row[5:]:
            if cell.value is None:
                break
            teachers.append(cell.value)

        lesson.append(teachers)
        lessons.append(lesson)

    return lessons

#"希望講師"ワークシートの情報を以下のような形式で保存する
# lessons = [
#               [学年, 生徒名, 科目名, コマ数, [希望講師1, 希望講師2, ...]], 
#               [学年, 生徒名, ...],
#               ... ,
#           ]
#
#
#
# BNF風に書くとこう
# lessons = [
#               [grade, name1, subject, num, teachers],
#               [grade, name2, subject, num, teachers],
#               ... ,
#           ]
# teachers = [teacher1, teacher2, ...]





def get_teachers(wb):
    
    ws_initial = wb["初期設定"]
    ws_schedule = wb["教室時間割"]

    teachers = []

    for initial_row in ws_initial.iter_rows(min_row=5, min_col=8):
        name = initial_row[0].value
        if name is None:
            continue
        teacher = [name]
        days = []
        # 同じ日の下のセルを見ないためのフラグ
        flag = 0

        for row in ws_schedule.iter_rows(min_row=4, min_col=2):

            # 無駄な範囲はbreak
            if row[0] is None:
                break

            # flagが1の時は下側を見ている
            if flag == 1:
                flag = 0
                continue

            # 今見てる名前でなければスキップ
            if row[1].value != name:
                continue
            # 該当氏名であればリストに情報を追加
            else:
                day = []

                for cell in row[4:23:3]:
                    if cell.fill == PatternFill(fill_type=None):
                        if cell.value is None:
                            day.append("free")
                        else:
                            day.append("attend")
                
                    else:
                        day.append("lock")

                flag = 1
                days.append(day)
        
        teacher.append(days)
        teachers.append(teacher)

    return teachers

    # 講師の出勤情報を以下の形式で保存
    # teachers = [
    #               [name, [[1日目の出勤情報（7要素配列）], [2日目], ...]],
    #               [name, [[1日目], ...]],
    #               ... ,
    #            ]
    # 出勤あり:attend, 白抜き:free, 出勤不可:lock






def student_free_count(students):
    free_counts = []
    for student in students:
        count = 0
        for day in student[1]:
            for lesson in day:
                if lesson == "free":
                    count += 1
        free_count = [student[0], count]
        free_counts.append(free_count)
    return free_counts

def free_count_bubble_sort(free_counts):
    length = len(free_counts)
    for i in range(0, length - 1):
        min = free_counts[i][1]
        indice = i
        for j in range(i + 1, length):
            if free_counts[j][1] < min:
                min = free_counts[j][1]
                indice = j
        free_counts[i], free_counts[indice] = free_counts[indice], free_counts[i]
    return free_counts