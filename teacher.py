import openpyxl as px
from openpyxl.styles import Font, Color, colors
from openpyxl.styles.fills import PatternFill

def teachers_attend_count(schedule):
    attend_counts = []
    for teacher in schedule:
        if teacher[0] is None:
            continue
        attend_count = [teacher[0]]
        count = 0
        for day in teacher[1]:
            for lesson in day:
                for koma in lesson:
                    if koma == ["free", "free", "free"]:
                        count += 1
        attend_count.append(count)
        attend_counts.append(attend_count)
    return attend_counts

def get_teachers(wb):
    
    ws_initial = wb["初期設定"]
    ws_schedule = wb["教室時間割"]

    teachers = []

    for initial_row in ws_initial.iter_rows(min_row=5, min_col=8):
        name = initial_row[0].value
        if name is None:
            continue
        teacher = [name]
        days = []
        # 同じ日の下のセルを見ないためのフラグ
        flag = 0

        for row in ws_schedule.iter_rows(min_row=4, min_col=2):

            # 無駄な範囲はbreak
            if row[0] is None:
                break

            # flagが1の時は下側を見ている
            if flag == 1:
                flag = 0
                continue

            # 今見てる名前でなければスキップ
            if row[1].value != name:
                continue
            # 該当氏名であればリストに情報を追加
            else:
                day = []

                for cell in row[4:23:3]:
                    if cell.fill == PatternFill(fill_type=None):
                        if cell.value is None:
                            day.append("free")
                        else:
                            day.append("attend")
                
                    else:
                        day.append("lock")

                flag = 1
                days.append(day)
        
        teacher.append(days)
        teachers.append(teacher)

    return teachers


    # 講師の出勤情報を以下の形式で保存
    # teachers = [
    #               [name, [[1日目の出勤情報（7要素配列）], [2日目], ...]],
    #               [name, [[1日目], ...]],
    #               ... ,
    #            ]
    # 出勤あり:attend, 白抜き:free, 出勤不可:lock